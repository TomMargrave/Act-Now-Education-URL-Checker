import csv
from bs4 import BeautifulSoup
import requests
import time
from requests.exceptions import RequestException, TooManyRedirects
from openpyxl import Workbook

def parse_table_to_csv(html_file, csv_file, skip_header=True):
    """
    Parses the first table in an HTML file and writes its contents to a CSV file.

    Args:
        html_file (str): Path to the input HTML file containing the table.
        csv_file (str): Path to the output CSV file.
        skip_header (bool): Whether to skip the first row (header) of the table. Defaults to True.

    Returns:
        bool: True if the table was found and written to CSV, False otherwise.
    """
    with open(html_file, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
    table = soup.find('table')
    if table is None:
        print("No table found in the HTML file.")
        return False
    rows = table.find_all('tr')
    start_idx = 1 if skip_header else 0
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['description', 'url', 'category'])
        writer.writeheader()
        for row in rows[start_idx:]:
            cols = row.find_all('td')
            if len(cols) >= 2:
                link = cols[0].find('a')
                if link and link.get('href'):
                    title = ' '.join(link.get_text(strip=True).split())
                    url = link['href']
                    category = cols[1].get_text(strip=True)
                    writer.writerow({'description': title, 'url': url, 'category': category})
    return True

def check_url_exists(url, allow_redirects=True, max_retries=2, retry_delay=1):
    """
    Checks if a given URL exists by sending a HEAD request, handling redirects and retries.

    Args:
        url (str): The URL to check.
        allow_redirects (bool, optional): Whether to follow redirects. Defaults to True.
        max_retries (int, optional): Maximum number of retry attempts on failure. Defaults to 2.
        retry_delay (int or float, optional): Delay in seconds between retries. Defaults to 1.

    Returns:
        tuple: A tuple containing:
            - status (str): One of "VALID", "INVALID", "REDIRECT_PERMANENT", "REDIRECT_TEMPORARY", "REDIRECT_OTHER", "REDIRECT_LOOP", or "ERROR".
            - status_code (int or str): HTTP status code or error message.
            - final_url (str): The final URL after redirects (or the original URL).
            - redirect_type (int or None): The HTTP status code of the redirect, if applicable.

    Raises:
        None: All exceptions are handled internally.
    """
    attempt = 0
    while attempt <= max_retries:
        try:
            response = requests.head(url, allow_redirects=allow_redirects, timeout=5)
            history = response.history
            if history:
                redirect_type = history[-1].status_code
                if redirect_type == 301:
                    status = "REDIRECT_PERMANENT"
                elif redirect_type == 302:
                    status = "REDIRECT_TEMPORARY"
                else:
                    status = "REDIRECT_OTHER"
                return status, response.status_code, response.url, redirect_type
            elif response.status_code == 200:
                return "VALID", 200, url, None
            else:
                return "INVALID", response.status_code, url, None
        except TooManyRedirects:
            return "REDIRECT_LOOP", 310, url, None
        except RequestException as e:
            if attempt == max_retries:
                return "ERROR", str(e), url, None
            else:
                time.sleep(retry_delay)
                attempt += 1

def validate_urls_from_csv(input_csv, output_all_xlsx, allow_redirects=True):
    """
    Validates URLs listed in a CSV file, checks their HTTP status, and writes the results to an Excel file.
    This function reads a CSV file containing URLs and their descriptions, validates each URL by checking its HTTP status,
    and categorizes them as valid, broken, skipped, or redirected. The results are written to an Excel workbook with
    separate sheets for all results, broken URLs, and valid URLs. It also prints a summary of the validation process.
    Args:
        input_csv (str): Path to the input CSV file. The CSV must have at least 'url' and 'description' columns.
        output_all_xlsx (str): Path to the output Excel (.xlsx) file where results will be saved.
        allow_redirects (bool, optional): Whether to allow HTTP redirects when checking URLs. Defaults to True.
    Returns:
        None
    Side Effects:
        - Writes an Excel file with validation results.
        - Prints progress and summary statistics to the console.
    Notes:
        - The function expects a helper function `check_url_exists(url, allow_redirects)` to be defined elsewhere.
        - The output Excel file will contain three sheets: "All URL Results", "Broken URLs", and "Valid URLs".
        - Special handling is performed for HTTP-to-HTTPS redirects and invalid HTTP URLs.
    """
    # Initialize counters for summary statistics
    total = valid = broken = skipped = redirected = 0
    invalid_http = 0
    redirect_permanent_http = 0

    # Create a new Excel workbook and sheets for all results, broken URLs, and valid URLs
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All URL Results"
    ws_broken = wb.create_sheet(title="Broken URLs")
    ws_valid = wb.create_sheet(title="Valid URLs")

    # Define headers for each sheet
    headers_all = ['Description', 'URL', 'Status', 'HTTP Status', 'Final URL', 'Redirect Type']
    ws_all.append(headers_all)
    ws_broken.append(headers_all)
    ws_valid.append(['Description', 'URL'])

    # Open the input CSV and process each row
    with open(input_csv, newline='', encoding='utf-8') as infile:
        reader = csv.DictReader(infile)
        for row_num, row in enumerate(reader, start=2):
            url = row.get('url', '').strip()
            description = row.get('description', '')
            if description is not None:
                description = description.strip().strip('"')
            else:
                description = ''
            total += 1

            if not url:
                # If URL is missing, skip and record as 'SKIPPED'
                status, code, final_url, redirect_type = 'SKIPPED', 'Missing URL', '', ''
                skipped += 1
                print(f"[{row_num}] SKIPPED     : (No URL)")
            else:
                # Check the URL using the helper function
                status, code, final_url, redirect_type = check_url_exists(url, allow_redirects)
                if status == "VALID":
                    valid += 1
                elif status.startswith("REDIRECT"):
                    redirected += 1
                else:
                    broken += 1

                # Special handling: If the URL is invalid but is an HTTP-to-HTTPS redirect, mark as INVALID_HTTP
                if (
                    status == "INVALID"
                    and url.startswith("http:")
                    and isinstance(final_url, str)
                    and final_url.startswith("https:")
                    and url.replace("http:", "https:", 1) == final_url
                ):
                    status = "INVALID_HTTP"

                # Adjust counters for INVALID_HTTP
                if status == "INVALID_HTTP":
                    broken -= 1
                    invalid_http += 1

                # Special handling: If the redirect is a permanent HTTP-to-HTTPS, mark as REDIRECT_PERMANENT_HTTP
                if (
                    status == "REDIRECT_PERMANENT"
                    and url.startswith("http:")
                    and isinstance(final_url, str)
                    and final_url.startswith("https:")
                    and url.replace("http:", "https:", 1) == final_url
                ):
                    status = "REDIRECT_PERMANENT_HTTP"

                # Adjust counters for REDIRECT_PERMANENT_HTTP
                if status == "REDIRECT_PERMANENT_HTTP":
                    redirected -= 1
                    redirect_permanent_http += 1

                print(f"[{row_num}] {status:<18}: {url}")

            # Write results to the "All URL Results" sheet
            ws_all.append([description, url, status, code, final_url, redirect_type])

            # Write valid URLs to the "Valid URLs" sheet
            if status == "VALID":
                ws_valid.append([description, url])

            # Write broken/error/redirect loop URLs to the "Broken URLs" sheet
            if status in ["INVALID", "ERROR", "REDIRECT_LOOP"]:
                ws_broken.append([description, url, status, code, final_url, redirect_type])

    # Save the Excel workbook with all results
    wb.save(output_all_xlsx)

    # Print summary statistics to the console
    print("\n📊 URL Validation Summary:")
    print(f"  Total URLs     : {total}")
    if valid != 0:
        print(f"  Valid URLs     : {valid}   {valid / total * 100:.2f}%")
    if redirected != 0:
        print(f"  Redirected     : {redirected}  {redirected / total * 100:.2f}%")
    if broken != 0:
        print(f"  Broken URLs    : {broken}   {broken / total * 100:.2f}%")
    if skipped != 0:
        print(f"  Skipped/Errors : {skipped}  {skipped / total * 100:.2f}%")
    if invalid_http != 0:
        print(f"  Invalid HTTP   : {invalid_http}  {invalid_http / total * 100:.2f}%")
    if redirect_permanent_http != 0:
        print(f"  Redirected HTTP: {redirect_permanent_http}  {redirect_permanent_http / total * 100:.2f}%")

    print(f"\n✅ Excel report written to: {output_all_xlsx}")

if __name__ == "__main__":
    html_file = 'element.html'
    csv_file = 'urls.csv'
    excel_file = 'url_validation_report.xlsx'

    if parse_table_to_csv(html_file, csv_file, skip_header=True):
        validate_urls_from_csv(
            input_csv=csv_file,
            output_all_xlsx=excel_file,
            allow_redirects=True
        )